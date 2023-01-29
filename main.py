from PIL import Image
from openpyxl import Workbook
from openpyxl import styles
# Open the image
im = Image.open("Landscape.png")

# Get the pixel data
pixels = im.load()
# Create a new Excel file
wb = Workbook()
ws = wb.active

# Iterate through the pixels
for i in range(im.width):
    for j in range(im.height):
        palette = im.getpalette()

        # Get the integer value of the pixel
        pixel_value = im.getpixel((i, j))

        # Get the RGB values of the pixel from the color palette
        red = palette[pixel_value * 3]
        green = palette[pixel_value * 3 + 1]
        blue = palette[pixel_value * 3 + 2]

        hex_color = '{:02x}{:02x}{:02x}'.format(red, green, blue)

        ws.cell(row=i + 1, column=j + 1).fill = styles.PatternFill(start_color=hex_color, end_color=hex_color,
                                                                            fill_type="solid")
# Save the Excel file
wb.save("pixel_art.xlsx")